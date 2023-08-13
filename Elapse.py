from genericpath import exists
import os,yaml,time
from pydoc import describe
import openpyxl
import argparse
from api.hostCollide import run_hostCollide
from api.fofa import fofa_search
from api.quake import quake_search
from api.zoomeye import zoomeye_search
from api.hunter import hunter_search
from api.proxy import check_proxy
from api.Aiqicha import run_aiqicha
from api.icp_select.icp_check import icp_check

parser = argparse.ArgumentParser(description="一键完成所有信息收集(被动)")
parser.add_argument('-d', '--domain', type=str, dest='domain', help='域名')
parser.add_argument('-c','--companyname', type=str,dest='companyname', help='公司名')
parser.add_argument('-df', '--domainfile', type=str, dest='domainfile', help='存放域名的文件')
parser.add_argument('-cf', '--companyfile', type=str, dest='companyfile', help='存放公司名的文件')
parser.add_argument('-p', '--proxy', dest='proxys', action="store_true", help='执行代理收集模块')

need_check_proxy={}

args = parser.parse_args()
py_path=os.path.dirname(os.path.realpath(__file__))         #获取py文件的当前路径
xlsx_save_name=str(py_path)+'/result/result-'+str(time.time_ns())+".xlsx"           #路径拼接、保存至result目录下

domain=args.domain          #获取域名
CompanyName=args.companyname            #获取公司名
domainfile=args.domainfile            #获取文件里的域名
companyfile=args.companyfile            #获取文件里的公司名
proxys=args.proxys          #如果用户未输入-p,则为false

if domain==None and CompanyName==None and domainfile==None and companyfile==None and proxys==None:          #公司名或者域名或者文件必须给一个
    print("[-] 未指定域名或公司名")
    exit()

def get_config():           #读取配置文件
    global py_path
    config_path=os.path.join(py_path,"config.yaml")         #路径拼接，获取py文件同目录下的config.yaml
    file = open(config_path,"r",encoding="utf-8")
    data = yaml.full_load(file)
    file.close()
    return data

def file_get(filename):
    targets=list()
    with open(filename,"r",encoding='utf-8') as fp:
        line=fp.readline()
        while line:
            line=line.strip()
            targets.append(line)
            line=fp.readline()
    return targets

all_config=get_config()         #获取所有密钥
excel=openpyxl.Workbook()           #每次运行都创建一个新的xlsx文件，而非覆盖｜追加
excel.remove(excel[excel.sheetnames[0]])


if domain:          #如果存在domain的值，则执行域名的空间聚合
    domains=domain.split(",")
elif domainfile:
    if exists(domainfile):            #如果存在文件，则domains的内容为文件内容（一行一个）
        domains=file_get(domainfile)
else:
    domains=None

if domains:
    check="select_domain"
    try:
        all_icp_domain=icp_check(domains,excel,xlsx_save_name)
        all_icp_domain=set(all_icp_domain)
        domains=set(domains)
        check_more_ip=all_icp_domain-domains
        if check_more_ip:
            need_append=str(input("icp备案查询中发现存在其他主域名，是否添加进列表(Y/N): "))
            if need_append.upper()=="Y":
                domains.update(all_icp_domain)
                print(domains)
    except Exception as e:
        print("icp备案模块存在问题")
        #print(e+"\n")
    domains=list(domains)

    for domain in domains:
        fofa_ips={}
        is_vip=all_config.get('fofa').get('vip')            #检查是否fofa的vip，默认不是
        if is_vip:
            fofa_data=["目标","ip","端口","title","host","域名","国家","省份","城市","国家全名","header","server","协议","banner","cert","运营商","as_number","as_organization","latitude","longitude","icp","fid","cname","type","jarm"]
        else:
            fofa_data=["目标","ip","端口","title","host","域名","国家","省份","城市","国家全名","header","server","协议","banner","cert","运营商","as_number","as_organization","latitude","longitude","icp","cname","type","jarm"]
        quake_data=["目标","ip","端口","title","域名","网页状态码","cert","传输协议类型","asn","org","国家","省份","城市","区域","框架详情","框架类型","框架等级","供应商","hostname","国家","省份","城市","区域","运营商","Body"]
        zoomeye_data=["目标","域名","时间戳","IP"]
        hunter_data=["目标","url","ip","域名","端口","网站标题","状态码","web服务","is_risk","is_risk_protocol","协议","基础协议","component","系统","公司","number","国家","省份","城市","更新时间","as_org","isp","banner"]

        # data=[]
        # for i in range(23):
        #     data.append(str(i))
        # w_excel=excel.create_sheet("api原始数据",0)
        # w_excel.append(data)            #添加title

        try:
            fofa_enable=all_config.get('fofa').get('enable')            #检查是否开启fofa模块
            if fofa_enable:
                fofa_ips=fofa_search(all_config,excel,fofa_data,xlsx_save_name,domain,check,is_vip)
            else:
                print("[-] fofa模块暂未开启...")
        except Exception as e:
            print("\n[-]fofa模块存在问题：")
            print(e)
            print("\n")
        try:
            quake_enable=all_config.get('quake').get('enable')          #检查是否开启quake模块
            if quake_enable:
                quake_search(all_config,excel,quake_data,xlsx_save_name,domain,check)
            else:
                print("[-] quake模块暂未开启...")
        except Exception as e:
            print("\n[-]quake模块存在问题：")
            print(e)
            print("\n")
        try:
            zoomeye_enable=all_config.get('zoomeye').get('enable')          #检查是否开启zoomeye模块
            if zoomeye_enable:
                zoomeye_search(all_config,excel,zoomeye_data,xlsx_save_name,domain)
            else:
                print("[-] zoomeye模块暂未开启...")
        except Exception as e:
            print("\n[-]zoomeye模块存在问题：")
            print(e)
            print("\n")
        try:
            hunter_enable=all_config.get('hunter').get('enable')          #检查是否开启hunter模块
            if hunter_enable:
                hunter_search(all_config,excel,hunter_data,xlsx_save_name,domain)
            else:
                print("[-] hunter模块暂未开启...")
        except Exception as e:
            print("\n[-]hunter模块存在问题：")
            print(e)
            print("\n")

        hostcollide_enable=all_config.get('hostcollide').get('enable')
        if hostcollide_enable:
            if fofa_ips:
                hostCollideResult, censysIPS = run_hostCollide(domain, fofa_ips, all_config)
            else:
                print("[-] 不进行host碰撞...")
        else:
            print("[-] host碰撞模块暂未开启...")

else:
    print("[-] 域名收集模块暂未开启...")
    
if CompanyName:         #如果存在CompanyName的值，则执行小程序、app的搜集模块
    companys=CompanyName.split(",")
elif companyfile:
    if exists(companyfile):
        companys=file_get(companyfile)
else:
    companys=None
if companys:
    for company in companys:
        aiqicha_enable=all_config.get('aiqicha').get('enable')
        if aiqicha_enable:
            selfIcpinfo_infos, invest_infos, holds_infos, branch_infos = run_aiqicha(excel,xlsx_save_name,company)
        
else:
    print("[-] 企业收集模块暂未开启。。。")

if proxys:          #如果Proxys的值为true，则执行代理收集模块
    proxy_enable=all_config.get('proxy').get('enable')            #检查是否开启proxy模块
    if proxy_enable:
        check="select_proxy"
        #fofa_proxy=fofa_search(all_config,excel,fofa_data,xlsx_save_name,domain,check)
        quake_proxy=quake_search(all_config,"","","",domain,check)
        fofa_proxy=fofa_search(all_config,"","","",domain,check)

        need_check_proxy.update(quake_proxy)
        need_check_proxy.update(fofa_proxy)
        #print(need_check_proxy)
        ok_proxy=check_proxy(need_check_proxy,py_path)
else:
    print("[-] 代理收集模块暂未开启...")
